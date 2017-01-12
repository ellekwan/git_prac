#!/usr/bin/python

import openpyxl, mysql, os, csv
from mysql.connector import MySQLConnection, Error
from openpyxl.utils import *
from print_table import print_tab

def open_exc():
##for testing multiple spreadsheets, delete later
    exc_files = {}
    count = 0
    try:
        for n in os.listdir('.'):
            if '.xlsx' in str(n):
                exc_files[count] = n
                count += 1
    except Error as e:
        print(e)
        print("No excel files found in CWD. Please change current working directory")
        exit(0)
    while True:
        try:
            print("Excel files in CWD:\n")
            for key, value in exc_files.items():
                print(key, ")   ", value)
            sel_exc = input("Select excel file to open\n type q to quit \n==> ")
            if sel_exc == 'q':
                print("Exiting...")
                exit(0)
            elif int(sel_exc) in exc_files.keys():
                sel_exc = int(sel_exc)
                global wkbk
                wkbk = exc_files[sel_exc]
                for name in (openpyxl.load_workbook(wkbk)).get_sheet_names():
                    if "Provision" not in name:
                        print("Invalid file, please choose a Provisioning form or quit and change directory.\n")
                        break
                    else:
                        print(wkbk, "is the Excel Document.\n")
                        global wb
                        wb = openpyxl.load_workbook(wkbk)
                        return wkbk, wb
                        break
            else:
                print("\n Please enter a valid option.\n")
                continue
        except ValueError:
            print("Invalid option, try again.")
            continue

##############################


def row_dict_values():
    """create a dictionary for each row"""
    sheet1 = wb.get_sheet_by_name("Operations Provisioning Form")
    col_max = get_column_letter(sheet1.max_column)
    row_max = sheet1.max_row
    cellmax = col_max + str(row_max)
    try:
        for row in sheet1["A1":cellmax]:
            for cell in row:
                if "NAME OF COMPANY" in str(cell.value):
                    co_cell = get_column_letter(column_index_from_string(cell.column) + 1) + str(cell.row)
                    global co
                    co = sheet1[co_cell].value
                    global row_dict
                    row_dict = {}
## for diff format provisioning sheets, shift cells vv
#                  for row in sheet1['A1':cellmax]:
#                      for cell in row:
#                          if "Username" in str(cell.value):
#                              cellrow = str(cell.row + 1)
#                              head_data = cell.column + cellrow
                    head_data = 'B7'
                    end_data = 'K' + str(row_max)
                    data_rng = sheet1[head_data:end_data]
                    for n in range(len(data_rng)):
                        row = data_rng[n]
                        name = row[0].value
## add in names to ignore
                        if name != None and \
                        "Steve Rogers" not in name and \
                        "Dave Holman" not in name and \
                        "Peter Jenkins" not in name and \
                        "Kayleigh McLintock" not in name and \
                        "Andy Lilly" not in name and \
                        "Echo Server" not in name and \
                        "Speaking Clock" not in name and \
                        "Template" not in name and \
                        "Conference Calling" not in name:
                            tel = row[3].value
                            cmty = row[9].value
                            email = row[4].value
                            make = row[1].value
                            model = row[2].value
                            if cmty == None:
                                cmty = row[5].value
                            row_dict.setdefault(name, [tel, cmty, None, None, None])
                            if tel == None or \
                                cmty == None or \
                                "-" in tel or \
                                "-" in cmty:
                                n += 1
                                input("Row for user ''{0}'' has incomplete user information. \
                                \nName, GSM number, and community are all required. \
                                \n User ''{1}'' will not be added to Database. \
                                \n\ncontinue:\n\n".format(name, name)) 
                            row_dict[name] = [tel, cmty, email, make, model]
                            n += 1

        print(co)
    except Error as error:
        print(error)
############################


def dict_tab_values():
    """insert name, tel, cmty, email, make, model into db and write to csv file"""

    csv_filename = wkbk.replace(".xlsx", ".csv")
    csvfile = open(csv_filename, 'w')
    wr = csv.writer(csvfile, quoting=csv.QUOTE_ALL)
##! import headers should be 'USER NAME', 'IDENTITY', 'EMAIL', 'PHONE NUMBER' 'PROVISION' according to csv
    for key, value in row_dict.items():
        name = key
        tel = value[0]
        cmty = value[1]
        email = value[2]
        make = value[3]
        model = value[4]
        print(name, value)
        if len(value) != 5:
            input("Error in row for user ''{0}''.\nUser ''{1}'' will not be added to Database".format(name, name))
            break
        cursorA = conn.cursor()
        cursorA.execute("CREATE DATABASE IF NOT EXISTS users_info_test;")
        cursorA.execute("USE users_info_test;")
        cursorA.execute("CREATE TABLE IF NOT EXISTS userInfo(company VARCHAR(50) NOT NULL, \
                                                             name VARCHAR(80) NOT NULL, \
                                                             gsm_tel VARCHAR(27) NOT NULL, \
                                                             community VARCHAR(99) NOT NULL, \
                                                             email VARCHAR(99) NULL, \
                                                             device_make VARCHAR(20) NULL, \
                                                             device_model VARCHAR(29) NULL, \
                                                             provision_date DATE NOT NULL, \
                                                             PRIMARY KEY(name, gsm_tel));")
        query_ins = """INSERT INTO userInfo VALUES("{0}", "{1}", "{2}", "{3}", "{4}", "{5}", "{6}", CURDATE());"""
        if tel != None and \
            cmty != None and \
            name != None and \
            "-" not in tel and \
            "-" not in cmty and \
            "N/A" not in tel and \
            "N/A" not in cmty:
            try:
                cursorA.execute(query_ins.format(co, name, tel, cmty, email, make, model))
                wr.writerow([name] + [tel] + [cmty] + [email] + [make] + [model])
                continue
            except Error as err:
                print(err)
                break 
        input("Error in row for user ''{0}''.\nUser ''{1}'' will not be added to Database".format(name, name))
    conn.commit()
    cursorA.close()
    print_tab(conn, "users_info_test", "userInfo")

#######################

def ex_main():
    open_exc()
    row_dict_values()
    dict_tab_values()

#######################

def __main__():
    try:
        print('Connecting to MySQL database...')
        global conn
        conn = MySQLConnection(user="guest", host="localhost", password="Armour_42")
        if conn.is_connected():
            print("Connection established.\n ")
        else: print("Connection failed. \n ")

    except Error as error:
        print(error)

    ex_main()
  
    while True:
        print("\n>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>\n\nLoad another provisioning form?\n\n0)     Yes\n1)     No\n2)     drop database")
        another = input("\nSelect an option\n==>\n")
        try:
            if int(another)==0:
                ex_main()
            elif int(another)==1:
                print("Closing Connection...")
                conn.close()
                exit(0)
####for testing multiple spreadsheets, delete later
            elif int(another)==2:
                cursorZ = conn.cursor()
                try:
                    print("dropping...")
                    cursorZ.execute("DROP DATABASE users_info_test;")
                    print("dropped.")
                    continue
                except Error as err:
                    print(err)
                    continue
                cursorZ.close()
####################
            else:
                print("Please enter a valid option.")
                continue
        except ValueError:
            print("Please enter a valid option")
            continue

__main__()
